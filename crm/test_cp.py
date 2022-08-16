from selenium import webdriver
from selenium.webdriver.support.ui import Select
import time
import openpyxl
import unicodedata

def findElements():
    #お客様情報
    elements_info = {
        "lastkana": driver.find_element_by_name("lastkana"), \
        "firstkana": driver.find_element_by_name("firstkana"), \
        "lastname": driver.find_element_by_name("lastname"), \
        "firstname": driver.find_element_by_name("firstname"), \
        "zipcode1": driver.find_element_by_name("zipcode1"), \
        "zipcode2": driver.find_element_by_name("zipcode2"), \
        "street": driver.find_element_by_name("street"), \
        "building": driver.find_element_by_name("building"), \
        "telno1": driver.find_element_by_name("telno1"), \
        "telno2": driver.find_element_by_name("telno2"), \
        "telno3": driver.find_element_by_name("telno3"), \
        "men": driver.find_element_by_xpath("//input[@name='sex'][@value='1']"), \
        "women": driver.find_element_by_xpath("//input[@name='sex'][@value='2']"), \
        "change_year": driver.find_element_by_id("change-year"), \
        "change_month": driver.find_element_by_id("change-month"), \
        "change_day": driver.find_element_by_id("change-day"), \
        "select_change_year": Select(change_year), \
        "select_change_month": Select(change_month), \
        "select_change_day": Select(change_day), \
        "mail": driver.find_element_by_name("mail"), \
        "mail2": driver.find_element_by_name("mail2"), \
        #imei　要素取得
        "imei": driver.find_element_by_name("imei"), \
        #クレジットカード情報　要素取得
        "card_no": driver.find_element_by_id("card_no"), \
        "card_expire_yyyy": driver.find_element_by_id("card_expire_yyyy"), \
        "card_expire_mm": driver.find_element_by_id("card_expire_mm"), \
        "select_card_expire_yyyy": Select(card_expire_yyyy), \
        "select_card_expire_mm": Select(card_expire_mm), \
        "card_lastkana": driver.find_element_by_id("card_lastkana"), \
        "card_firstkana": driver.find_element_by_id("card_firstkana"), \
        "card_birth_date_yyyy": driver.find_element_by_id("card_birth_date_yyyy"), \
        "card_birth_date_mm": driver.find_element_by_id("card_birth_date_mm"), \
        "card_birth_date_dd": driver.find_element_by_name("card_birth_date_dd"), \
        "select_card_birth_date_yyyy": Select(card_birth_date_yyyy), \
        "select_card_birth_date_mm": Select(card_birth_date_mm), \
        "select_card_birth_date_dd": Select(card_birth_date_dd), \
        "security_cd": driver.find_element_by_id("security_cd"), \
        "confire_btn": driver.find_element_by_id("confire_btn")
    }
    return elements_info

driver = webdriver.Chrome("C:\\Users\\kyosuke.ushiyama\\AppData\\Local\\SeleniumBasic\\chromedriver.exe")
driver.get("http://entrylocal.examobile.jp/")
image_file = ("C:\\Users\\kyosuke.ushiyama\\python\\license.jpg")

book = openpyxl.load_workbook("C:\\Users\\kyosuke.ushiyama\\python\\customer_info.xlsx", data_only=True)
sheet = book["お客様情報"]
#print(sheet["C1"].value)
num_of_customers = int(sheet["C1"].value)

start_row = 6
start_column = 3


for i in range(num_of_customers):
    driver.get("http://entrylocal.examobile.jp/")
    customer_info = {"lastkana": sheet.cell(row = start_row, column = start_column).value, \
                    "firstkana": sheet.cell(row = start_row + 1, column = start_column).value, \
                    "lastname": sheet.cell(row = start_row + 2, column = start_column).value, \
                    "firstname": sheet.cell(row = start_row + 3, column = start_column).value, \
                    "zipcode": sheet.cell(row = start_row + 4, column = start_column).value, \
                    "street": sheet.cell(row = start_row + 5, column = start_column).value, \
                    "building": sheet.cell(row = start_row + 6, column = start_column).value, \
                    "telno": sheet.cell(row = start_row + 7, column = start_column).value, \
                    "sex": sheet.cell(row = start_row + 8, column = start_column).value, \
                    "birth": sheet.cell(row = start_row + 9, column = start_column).value, \
                    "mail": sheet.cell(row = start_row + 10, column = start_column).value, \
                    "imei": sheet.cell(row = start_row + 11, column = start_column).value, \
                    "card_no": sheet.cell(row = start_row + 12, column = start_column).value, \
                    "card_expire": sheet.cell(row = start_row + 13, column = start_column).value, \
                    "card_lastkana": sheet.cell(row = start_row + 14, column = start_column).value, \
                    "card_firstkana": sheet.cell(row = start_row + 15, column = start_column).value, \
                    "card_birth": sheet.cell(row = start_row + 16, column = start_column).value, \
                    "security_cd": sheet.cell(row = start_row + 17, column = start_column).value
                    }

    for key in customer_info:
        if customer_info[key] is None:
            customer_info[key] = ""

    #同意
    check = driver.find_element_by_xpath("//input[@id='consent']")
    check.click()
    time.sleep(1)
    entry = driver.find_element_by_css_selector(".btn.btn-primary")
    entry.click()

    time.sleep(2)
    
    # 要素取得
    elements = findElements()
    print(elements)

    #お客様情報入力
    lastkana.send_keys(customer_info["lastkana"])
    firstkana.send_keys(customer_info["firstkana"])
    lastname.send_keys(customer_info["lastname"])
    firstname.send_keys(customer_info["firstname"])

    str_zipcode = customer_info["zipcode"].split("-")
    zipcode1.send_keys(str_zipcode[0])
    zipcode2.send_keys(str_zipcode[1])
    driver.execute_script("search_address()")
    time.sleep(3)
    street.send_keys(customer_info["street"])
    building.send_keys(customer_info["building"])

    str_telno = customer_info["telno"].split("-")
    telno1.send_keys(str_telno[0])
    telno2.send_keys(str_telno[1])
    telno3.send_keys(str_telno[2])

    if customer_info["sex"] == "男性":
        #男性の場合
        men.click()
    elif customer_info["sex"] == "女性":
        #女性の場合
        women.click()

    str_birth = customer_info["birth"].split("/")
    select_change_year.select_by_value(str_birth[0])
    select_change_month.select_by_value(str_birth[1])
    select_change_day.select_by_value(str_birth[2])

    mail.send_keys(customer_info["mail"])
    mail2.send_keys(customer_info["mail"])
    imei.send_keys(customer_info["imei"])
    card_no.send_keys(customer_info["card_no"])

    str_card_expire = customer_info["card_expire"].split("/")
    select_card_expire_yyyy.select_by_value(str_card_expire[0])
    select_card_expire_mm.select_by_value(str_card_expire[1])
    card_lastkana.send_keys(unicodedata.normalize("NFKC", customer_info["card_lastkana"]))
    card_firstkana.send_keys(unicodedata.normalize("NFKC", customer_info["card_firstkana"]))
    str_card_birth = customer_info["card_birth"].split("/")
    select_card_birth_date_yyyy.select_by_value(str_card_birth[0])
    select_card_birth_date_mm.select_by_value(str_card_birth[1])
    select_card_birth_date_dd.select_by_value(str_card_birth[2])
    security_cd.send_keys(customer_info["security_cd"])
    #confire_btn.click()

    # /apply/inputページ
    time.sleep(2)
    apply = driver.find_element_by_css_selector(".btn.btn-primary")
    #apply.click()

    # /apply/identifyページ
    time.sleep(2)
    main_identity = driver.find_element_by_xpath("//input[@name='main_identity'][@value='1']")
    main_identity.click()
    image1 = driver.find_element_by_name("image1")
    image2 = driver.find_element_by_name("image2")
    image1.send_keys(image_file)
    image2.send_keys(image_file)
    image4 = driver.find_element_by_name("image4")
    image4.send_keys(image_file)
    complete_btn = driver.find_element_by_id("complete_btn")
    complete_btn.click()

    time.sleep(2)

    start_column = start_column + 1




lastkana = driver.find_element_by_name("lastkana")
    firstkana = driver.find_element_by_name("firstkana")
    lastname = driver.find_element_by_name("lastname")
    firstname = driver.find_element_by_name("firstname")
    zipcode1 = driver.find_element_by_name("zipcode1")
    zipcode2 = driver.find_element_by_name("zipcode2")
    street = driver.find_element_by_name("street")
    building = driver.find_element_by_name("building")
    telno1 = driver.find_element_by_name("telno1")
    telno2 = driver.find_element_by_name("telno2")
    telno3 = driver.find_element_by_name("telno3")
    men = driver.find_element_by_xpath("//input[@name='sex'][@value='1']")
    women = driver.find_element_by_xpath("//input[@name='sex'][@value='2']")
    change_year = driver.find_element_by_id("change-year")
    change_month = driver.find_element_by_id("change-month")
    change_day = driver.find_element_by_id("change-day")
    select_change_year = Select(change_year)
    select_change_month = Select(change_month)
    select_change_day = Select(change_day)
    mail = driver.find_element_by_name("mail")
    mail2 = driver.find_element_by_name("mail2")
    #imei　要素取得
    imei = driver.find_element_by_name("imei")
    #クレジットカード情報　要素取得
    card_no = driver.find_element_by_id("card_no")
    card_expire_yyyy = driver.find_element_by_id("card_expire_yyyy")
    card_expire_mm = driver.find_element_by_id("card_expire_mm")
    select_card_expire_yyyy = Select(card_expire_yyyy)
    select_card_expire_mm = Select(card_expire_mm)
    card_lastkana = driver.find_element_by_id("card_lastkana")
    card_firstkana = driver.find_element_by_id("card_firstkana")
    card_birth_date_yyyy = driver.find_element_by_id("card_birth_date_yyyy")
    card_birth_date_mm = driver.find_element_by_id("card_birth_date_mm")
    card_birth_date_dd = driver.find_element_by_name("card_birth_date_dd")
    select_card_birth_date_yyyy = Select(card_birth_date_yyyy)
    select_card_birth_date_mm = Select(card_birth_date_mm)
    select_card_birth_date_dd = Select(card_birth_date_dd)
    security_cd = driver.find_element_by_id("security_cd")
    confire_btn = driver.find_element_by_id("confire_btn")
