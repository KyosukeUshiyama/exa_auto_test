from selenium import webdriver
from selenium.webdriver.support.ui import Select
import time
import unicodedata
import openpyxl


#エクセルより顧客情報読み取り
def getCustomerInfo(start_row, start_column, sheet):
    kana = sheet.cell(row = start_row, column = start_column).value
    name = sheet.cell(row = start_row + 1, column = start_column).value
    street_building = sheet.cell(row = start_row + 3, column = start_column).value
    if(len(street_building.split()) == 1):
        street = street_building
        building = ""
    else:
        street = street_building.split()[0]
        building = street_building.split()[1]
    cardkana = sheet.cell(row = start_row + 11, column = start_column).value
    
    customer_info = {
        "lastkana": kana.split()[0], \
        "firstkana": kana.split()[1], \
        "lastname": name.split()[0], \
        "firstname": name.split()[1], \
        "zipcode": sheet.cell(row = start_row + 2, column = start_column).value, \
        "street": street, \
        "building": building, \
        "telno": sheet.cell(row = start_row + 4, column = start_column).value, \
        "sex": sheet.cell(row = start_row + 5, column = start_column).value, \
        "birth": sheet.cell(row = start_row + 6, column = start_column).value, \
        "mail": sheet.cell(row = start_row + 7, column = start_column).value, \
        "imei": sheet.cell(row = start_row + 8, column = start_column).value, \
        "card_no": sheet.cell(row = start_row + 9, column = start_column).value, \
        "card_expire": sheet.cell(row = start_row + 10, column = start_column).value, \
        "card_lastkana": cardkana.split()[0], \
        "card_firstkana": cardkana.split()[1], \
        "card_birth": sheet.cell(row = start_row + 12, column = start_column).value, \
        "security_cd": sheet.cell(row = start_row + 13, column = start_column).value
    }
    return customer_info

#ページ内要素取得#
def getElementsInfo(driver):
    elements_info = {
        "lastkana": driver.find_element_by_name("lastkana"), \
        "firstkana": driver.find_element_by_name("firstkana"), \
        "lastname": driver.find_element_by_name("lastname"), \
        "firstname": driver.find_element_by_name("firstname"), \
        "zipcode1": driver.find_element_by_name("zipcode1"), \
        "zipcode2": driver.find_element_by_name("zipcode2"), \
        "street": driver.find_element_by_name("street"), \
        "building": driver.find_element_by_name("building"), \
        "telno1": driver.find_element_by_name("telno1"), \
        "telno2": driver.find_element_by_name("telno2"), \
        "telno3": driver.find_element_by_name("telno3"), \
        "men": driver.find_element_by_xpath("//input[@name='sex'][@value='1']"), \
        "women": driver.find_element_by_xpath("//input[@name='sex'][@value='2']"), \
        "change_year": Select(driver.find_element_by_id("change-year")), \
        "change_month": Select(driver.find_element_by_id("change-month")), \
        "change_day": Select(driver.find_element_by_id("change-day")), \
        #"select_change_year": Select(change_year), \
        #"select_change_month": Select(change_month), \
        #"select_change_day": Select(change_day), \
        "mail": driver.find_element_by_name("mail"), \
        "mail2": driver.find_element_by_name("mail2"), \
        #imei　要素取得
        "imei": driver.find_element_by_name("imei"), \
        #クレジットカード情報　要素取得
        "card_no": driver.find_element_by_id("card_no"), \
        "card_expire_yyyy": Select(driver.find_element_by_id("card_expire_yyyy")), \
        "card_expire_mm": Select(driver.find_element_by_id("card_expire_mm")), \
        #"select_card_expire_yyyy": Select(card_expire_yyyy), \
        #"select_card_expire_mm": Select(card_expire_mm), \
        "card_lastkana": driver.find_element_by_id("card_lastkana"), \
        "card_firstkana": driver.find_element_by_id("card_firstkana"), \
        "card_birth_date_yyyy": Select(driver.find_element_by_id("card_birth_date_yyyy")), \
        "card_birth_date_mm": Select(driver.find_element_by_id("card_birth_date_mm")), \
        "card_birth_date_dd": Select(driver.find_element_by_name("card_birth_date_dd")), \
        #"select_card_birth_date_yyyy": Select(card_birth_date_yyyy), \
        #"select_card_birth_date_mm": Select(card_birth_date_mm), \
        #"select_card_birth_date_dd": Select(card_birth_date_dd), \
        "security_cd": driver.find_element_by_id("security_cd"), \
        "confire_btn": driver.find_element_by_id("confire_btn")
    }
    return elements_info


#お客様情報入力
def insertCustomerInfo(elements_info, customer_info, driver):
    elements_info["lastkana"].send_keys(customer_info["lastkana"])
    elements_info["firstkana"].send_keys(customer_info["firstkana"])
    elements_info["lastname"].send_keys(customer_info["lastname"])
    elements_info["firstname"].send_keys(customer_info["firstname"])

    str_zipcode = customer_info["zipcode"].split("-")
    elements_info["zipcode1"].send_keys(str_zipcode[0])
    elements_info["zipcode2"].send_keys(str_zipcode[1])
    driver.execute_script("search_address()")
    time.sleep(3)
    elements_info["street"].send_keys(customer_info["street"])
    elements_info["building"].send_keys(customer_info["building"])

    str_telno = customer_info["telno"].split("-")
    elements_info["telno1"].send_keys(str_telno[0])
    elements_info["telno2"].send_keys(str_telno[1])
    elements_info["telno3"].send_keys(str_telno[2])

    if customer_info["sex"] == "男性" or customer_info["sex"] == "男":
        #男性の場合
        elements_info["men"].click()
    elif customer_info["sex"] == "女性" or customer_info["sex"] == "女":
        #女性の場合
        elements_info["women"].click()

    str_birth = customer_info["birth"].split("/")
    elements_info["change_year"].select_by_value(str_birth[0])
    if(len(str_birth[1]) == 1):
        str_birth[1] = "0" + str_birth[1]
    elements_info["change_month"].select_by_value(str_birth[1])
    if(len(str_birth[2]) == 1):
        str_birth[2] = "0" + str_birth[2]
    elements_info["change_day"].select_by_value(str_birth[2])
    elements_info["mail"].send_keys(customer_info["mail"])
    elements_info["mail2"].send_keys(customer_info["mail"])

    ##########
    #imei入力#
    ##########
    elements_info["imei"].send_keys(customer_info["imei"])
    
    ########################
    #クレジットカード情報入力#
    ########################
    elements_info["card_no"].send_keys(customer_info["card_no"])
    str_card_expire = customer_info["card_expire"].split("/")
    elements_info["card_expire_yyyy"].select_by_value(str_card_expire[0])
    elements_info["card_expire_mm"].select_by_value(str_card_expire[1])
    elements_info["card_lastkana"].send_keys(unicodedata.normalize("NFKC", customer_info["card_lastkana"]))
    elements_info["card_firstkana"].send_keys(unicodedata.normalize("NFKC", customer_info["card_firstkana"]))
    str_card_birth = customer_info["card_birth"].split("/")
    elements_info["card_birth_date_yyyy"].select_by_value(str_card_birth[0])
    if(len(str_card_birth[1]) == 1):
        str_card_birth[1] = "0" + str_card_birth[1]
    elements_info["card_birth_date_mm"].select_by_value(str_card_birth[1])
    if(len(str_card_birth[2]) == 1):
        str_card_birth[2] = "0" + str_card_birth[2]
    elements_info["card_birth_date_dd"].select_by_value(str_card_birth[2])
    elements_info["security_cd"].send_keys(customer_info["security_cd"])
